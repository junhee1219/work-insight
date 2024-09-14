from openpyxl import load_workbook
import os

def get_filename_from_path(filepath):
    basename = os.path.basename(filepath)  
    return os.path.splitext(basename)[0]

def get_sequences(list_of_ints):
    sequence_count = 1
    sequences = []
    for row in list_of_ints:
        next_item = None
        if list_of_ints.index(row) < (len(list_of_ints) - 1):
            next_item = list_of_ints[list_of_ints.index(row) + 1]
        if (row + 1) == next_item:
            sequence_count += 1
        else:
            first_in_sequence = list_of_ints[list_of_ints.index(row) - sequence_count + 1]
            sequences.append([first_in_sequence, sequence_count])
            sequence_count = 1
    return sequences

def get_delete_row(sheet, target_column, filtered_data, max_row, start_row):
# 삭제할 행 번호를 저장할 리스트
    rows_to_delete = []
    # 2행부터 마지막 행까지 반복하면서 B열의 값을 확인
    for row in range(start_row, max_row + 1):
        if sheet[f'{target_column}{row}'].value != filtered_data:
            rows_to_delete.append(row)
    return get_sequences(rows_to_delete)

def split_main(filename, target_column_list, start_row_list):
    workbook = load_workbook(filename, data_only=True)
    sheet_list = workbook.sheetnames
    split_data = set()
    for split_name in sheet_list:
        sheet = workbook[split_name]
        max_row = sheet.max_row
        target_column = target_column_list.get(split_name, "None")
        if target_column == "None":
            continue
        for row in range(start_row_list[split_name], max_row + 1):
            split_data.add(sheet[f'{target_column}{row}'].value)  ## {target_column}{row} : A1 과 같은 형태
        
    # split_data로부터 각 파일 생성 , 분리
    for split_name in split_data:
        filepath = os.path.dirname(filename) + f"/{get_filename_from_path(filename).replace("/","")}_{split_name.replace("/","")}.xlsx"
        workbook.save(filepath)
        this_workbook = load_workbook(filepath, data_only=True)

        for sheet_name in sheet_list:
            cur_sheet = this_workbook[sheet_name]
            target_column = target_column_list.get(sheet_name, "None")
            max_row = cur_sheet.max_row
            start_row = start_row_list[sheet_name]
            if target_column == "None":
                continue
            for sequence in reversed(get_delete_row(cur_sheet, target_column, split_name, max_row, start_row)):
                cur_sheet.delete_rows(sequence[0],sequence[1])
        this_workbook.save(filepath)
